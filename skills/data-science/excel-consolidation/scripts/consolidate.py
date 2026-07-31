"""
Generic multi-sheet Excel consolidator.
Usage: python consolidate.py <input.xlsx> [output.xlsx]

Reads all sheets, normalizes columns to the widest format,
deduplicates exact row matches, writes single-sheet output.

Modify COLUMN_MAP for each sheet that has a different column layout.
"""
import sys, os
from openpyxl import Workbook, load_workbook

def build_col_map(sheet_name, src_cols, target_cols):
    """
    Return {src_idx: target_idx} mapping.
    Override this for sheets with different column layouts.
    If src_cols == target_cols (same length + order), returns identity map.
    """
    if len(src_cols) == len(target_cols):
        return {i: i for i in range(len(src_cols))}
    # Custom per-sheet mappings go here
    # Example: ANTICORRUPCIÓN (9 cols) → target (11 cols)
    # missing cols 4 (FECHA DE INICIO) and 9 (MATERIA)
    CUSTOM_MAPS = {
        'ANTICORRUPCIÓN': {0: 0, 1: 1, 2: 2, 3: 3, 4: 5, 5: 6, 6: 7, 7: 8, 8: 10},
    }
    return CUSTOM_MAPS.get(sheet_name, {i: i for i in range(min(len(src_cols), len(target_cols)))})

def normalize(values, col_map, n_target):
    row = [None] * n_target
    for src, dst in col_map.items():
        if src < len(values):
            row[dst] = values[src]
    return row

def row_key(row):
    return tuple(str(v).strip().lower() if v is not None else '' for v in row)

def main():
    if len(sys.argv) < 2:
        print(f'Usage: {sys.argv[0]} <input.xlsx> [output.xlsx]')
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else (
        os.path.join(os.path.dirname(input_path),
                     f'CONSOLIDADO_{os.path.basename(input_path)}')
    )

    wb_in = load_workbook(input_path, read_only=True)
    sheets = wb_in.sheetnames

    # Determine target columns from the sheet with most columns
    # (skip title rows 0-2, read header row 3)
    target_cols = []
    for sname in sheets:
        ws = wb_in[sname]
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=4)):
            if i == 2:  # header row
                header = [str(v).strip() if v else '' for v in row]
                if len(header) > len(target_cols):
                    target_cols = header
                break
    if not target_cols:
        print('ERROR: could not determine target columns')
        sys.exit(1)

    print(f'Target columns ({len(target_cols)}): {target_cols}')
    print(f'Sheets: {sheets}')

    seen = set()
    all_rows = []
    total_read = 0

    for sname in sheets:
        ws = wb_in[sname]
        col_map = build_col_map(sname, target_cols, target_cols)
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3:
                continue
            vals = list(row)
            if all(v is None for v in vals):
                continue
            norm = normalize(vals, col_map, len(target_cols))
            if not norm[3]:  # skip if key col (index 3) is empty
                continue
            key = row_key(norm)
            if key not in seen:
                seen.add(key)
                all_rows.append(norm)
            count += 1
        print(f'  {sname}: {count} rows → {len(all_rows) - total_read} new unique')
        total_read += count
    wb_in.close()

    print(f'\nTotal: {total_read} source → {len(all_rows)} unique')

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = 'CONSOLIDADO'
    ws_out.append(target_cols)

    BATCH = 10000
    for i in range(0, len(all_rows), BATCH):
        for row in all_rows[i:i + BATCH]:
            ws_out.append(row)
        print(f'  Written {min(i + BATCH, len(all_rows))}/{len(all_rows)}')

    wb_out.save(output_path)
    print(f'\n✅ Created: {output_path} ({len(all_rows)} rows)')

if __name__ == '__main__':
    main()
